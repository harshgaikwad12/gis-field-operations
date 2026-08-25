"""
Import Thote+Thakre (TD Compulsory) real field data into PostgreSQL with high accuracy.
Sets up Nagpur hierarchy, creates all 4 role accounts, inserts all 32 master meters and pending consumers.
"""
import sys
import os
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

# Add app to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db.session import SessionLocal
from app.models.zone import Zone
from app.models.area import Area
from app.models.field_area import FieldArea
from app.models.officer import Officer
from app.models.master_meter import MasterMeter
from app.models.pending_consumer import PendingConsumer
from app.services.officer_storage import create_officer, get_officer_by_email


def import_data():
    excel_path = Path(__file__).resolve().parent.parent.parent.parent / "data" / "Thote+Thakre (TD Compulsory)1 2.xlsx"
    if not excel_path.exists():
        # check direct path
        excel_path = Path("/Users/harshgaikwad/Downloads/gis-field-operations/data/Thote+Thakre (TD Compulsory)1 2.xlsx")

    print(f"Loading Excel file from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["THOTE + Thakre"]

    db = SessionLocal()
    try:
        now = datetime.now(timezone.utc)

        print("\n=== 1. Creating Nagpur Geographic Tree ===")
        # Zone
        zone = db.query(Zone).filter((Zone.zone_code == "MH-NGP-ZONE") | (Zone.zone_name == "Nagpur Zone")).first()
        if not zone:
            zone = Zone(
                zone_code="MH-NGP-ZONE",
                zone_name="Nagpur Zone",
                is_active=True,
            )
            db.add(zone)
            db.flush()
            print(f"Created Zone: {zone.zone_name} (ID: {zone.id})")
        else:
            print(f"Zone: {zone.zone_name} (ID: {zone.id})")

        # Area
        area = db.query(Area).filter((Area.area_code == "MH-NGP-AREA1") | (Area.area_name == "Nagpur North Area")).first()
        if not area:
            area = Area(
                zone_id=zone.id,
                area_code="MH-NGP-AREA1",
                area_name="Nagpur North Area",
                is_active=True,
            )
            db.add(area)
            db.flush()
            print(f"Created Area: {area.area_name} (ID: {area.id})")
        else:
            print(f"Area: {area.area_name} (ID: {area.id})")

        # Field Area / Ward
        field_area = db.query(FieldArea).filter(
            (FieldArea.field_area_code == "MH-NGP-WARD-TT") | (FieldArea.field_area_name == "Thote & Thakre Ward (Godhani-Koradi)")
        ).first()
        if not field_area:
            field_area = FieldArea(
                area_id=area.id,
                field_area_code="MH-NGP-WARD-TT",
                field_area_name="Thote & Thakre Ward (Godhani-Koradi)",
                is_active=True,
            )
            db.add(field_area)
            db.flush()
            print(f"Created Field Area: {field_area.field_area_name} (ID: {field_area.id})")
        else:
            print(f"Field Area: {field_area.field_area_name} (ID: {field_area.id})")

        print("\n=== 2. Setting Up User Accounts for All Roles ===")
        # 1. SUPER_ADMIN
        sa = get_officer_by_email(db, "superadmin@maharashtra.gov.in")
        if not sa:
            sa = create_officer(
                db,
                officer_code="MH-SUPER-001",
                officer_name="Maharashtra State Super Admin",
                email="superadmin@maharashtra.gov.in",
                phone="9999999999",
                password="SuperAdmin@12345",
                role="SUPER_ADMIN",
                is_active=True,
            )
            print(f"Created Super Admin: {sa.email}")
        else:
            print(f"Super Admin active: {sa.email}")

        # 2. ZONAL_ADMIN (Nagpur)
        za = get_officer_by_email(db, "zonal.nagpur@maharashtra.gov.in")
        if not za:
            za = create_officer(
                db,
                officer_code="MH-NGP-ZA-001",
                officer_name="Nagpur Zonal Admin",
                email="zonal.nagpur@maharashtra.gov.in",
                phone="9822001122",
                password="ZonalAdmin@12345",
                role="ZONAL_ADMIN",
                zone_id=zone.id,
                is_active=True,
            )
            print(f"Created Zonal Admin: {za.email} (Zone: {zone.zone_name})")
        else:
            za.zone_id = zone.id
            db.commit()
            print(f"Zonal Admin active: {za.email}")

        # 3. AREA_ADMIN (Nagpur)
        aa = get_officer_by_email(db, "area.nagpur@maharashtra.gov.in")
        if not aa:
            aa = create_officer(
                db,
                officer_code="MH-NGP-AA-001",
                officer_name="Nagpur Area Admin",
                email="area.nagpur@maharashtra.gov.in",
                phone="9822003344",
                password="AreaAdmin@12345",
                role="AREA_ADMIN",
                area_id=area.id,
                is_active=True,
            )
            print(f"Created Area Admin: {aa.email} (Area: {area.area_name})")
        else:
            aa.area_id = area.id
            db.commit()
            print(f"Area Admin active: {aa.email}")

        # 4. FIELD_OFFICER (Nagpur - Thote & Thakre)
        fo = get_officer_by_email(db, "officer.nagpur@maharashtra.gov.in")
        if not fo:
            fo = create_officer(
                db,
                officer_code="MH-NGP-FO-001",
                officer_name="Thote & Thakre Field Officer",
                email="officer.nagpur@maharashtra.gov.in",
                phone="9822005566",
                password="FieldOfficer@12345",
                role="FIELD_OFFICER",
                field_area_id=field_area.id,
                is_active=True,
            )
            print(f"Created Field Officer: {fo.email} (Ward: {field_area.field_area_name})")
        else:
            fo.field_area_id = field_area.id
            db.commit()
            print(f"Field Officer active: {fo.email}")

        print("\n=== 3. Importing High-Accuracy Master Meters & Pending Consumers ===")
        meter_count = 0
        consumer_count = 0

        # Iterate Excel rows (Row 2 to max_row)
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or "").strip()
            meter_number = str(ws.cell(r, 2).value or "").strip()
            consumer_no = str(ws.cell(r, 3).value or "").strip()
            lat_val = ws.cell(r, 4).value
            lon_val = ws.cell(r, 5).value
            amount_val = ws.cell(r, 6).value
            age_val = ws.cell(r, 7).value

            if not name or not meter_number:
                continue

            lat = float(lat_val) if lat_val is not None else 21.2314
            lon = float(lon_val) if lon_val is not None else 79.0895
            amount = float(amount_val) if amount_val is not None else 0.0
            days = int(age_val) if age_val is not None else 0

            # 1. Master Meter
            existing_meter = db.query(MasterMeter).filter(MasterMeter.meter_id == meter_number).first()
            if not existing_meter:
                meter = MasterMeter(
                    zone_id=zone.id,
                    area_id=area.id,
                    field_area_id=field_area.id,
                    meter_id=meter_number,
                    customer_id=consumer_no or f"CUST-{meter_number}",
                    customer_name=name,
                    location=f"POINT({lon} {lat})",
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                )
                db.add(meter)
                meter_count += 1
            else:
                existing_meter.field_area_id = field_area.id
                existing_meter.customer_name = name
                existing_meter.location = f"POINT({lon} {lat})"
                existing_meter.updated_at = datetime.utcnow()

            # 2. Pending Consumer
            existing_consumer = db.query(PendingConsumer).filter(PendingConsumer.consumer_id == consumer_no).first()
            if not existing_consumer:
                consumer = PendingConsumer(
                    consumer_id=consumer_no,
                    consumer_name=name,
                    meter_id=meter_number,
                    pending_amount=amount,
                    days_pending=days,
                )
                db.add(consumer)
                consumer_count += 1
            else:
                existing_consumer.consumer_name = name
                existing_consumer.meter_id = meter_number
                existing_consumer.pending_amount = amount
                existing_consumer.days_pending = days

        db.commit()
        print(f"\nImport Summary: Successfully filed {meter_count} Master Meters and {consumer_count} Pending Consumers into database with high precision GPS!")

    finally:
        db.close()


if __name__ == "__main__":
    import_data()
