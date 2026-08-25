import csv
import io
from dataclasses import dataclass

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "meter_id",
    "customer_id",
    "customer_name",
    "latitude",
    "longitude",
)


@dataclass
class ValidationError:
    row: int
    field: str
    message: str


def read_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")

    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        raise ValueError("File does not contain a header row.")

    return [
        {
            key.strip(): (value or "").strip()
            for key, value in row.items()
            if key is not None
        }
        for row in reader
    ]


def read_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("File does not contain a header row.") from exc

    headers = [
        str(value).strip() if value is not None else ""
        for value in header_row
    ]

    return [
        {
            headers[index]: (
                str(value).strip()
                if value is not None
                else ""
            )
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        for row in rows
    ]


def validate_columns(rows: list[dict[str, str]]) -> list[ValidationError]:
    if not rows:
        return [
            ValidationError(
                row=1,
                field="file",
                message="File contains no data rows.",
            )
        ]

    columns = set(rows[0].keys())

    return [
        ValidationError(
            row=1,
            field=column,
            message="Required column is missing.",
        )
        for column in REQUIRED_COLUMNS
        if column not in columns
    ]


def validate_rows(
    rows: list[dict[str, str]],
) -> list[ValidationError]:
    errors: list[ValidationError] = []
    meter_ids: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        for field in REQUIRED_COLUMNS:
            if not row.get(field, "").strip():
                errors.append(
                    ValidationError(
                        row=row_number,
                        field=field,
                        message="Value is required.",
                    )
                )

        meter_id = row.get("meter_id", "").strip()

        if meter_id:
            if meter_id in meter_ids:
                errors.append(
                    ValidationError(
                        row=row_number,
                        field="meter_id",
                        message="Duplicate Meter ID.",
                    )
                )
            else:
                meter_ids.add(meter_id)

        latitude = row.get("latitude", "").strip()

        if latitude:
            try:
                value = float(latitude)

                if not -90 <= value <= 90:
                    raise ValueError

            except ValueError:
                errors.append(
                    ValidationError(
                        row=row_number,
                        field="latitude",
                        message="Latitude must be between -90 and 90.",
                    )
                )

        longitude = row.get("longitude", "").strip()

        if longitude:
            try:
                value = float(longitude)

                if not -180 <= value <= 180:
                    raise ValueError

            except ValueError:
                errors.append(
                    ValidationError(
                        row=row_number,
                        field="longitude",
                        message="Longitude must be between -180 and 180.",
                    )
                )

    return errors


def validate_master_meter_file(
    filename: str,
    content: bytes,
) -> tuple[list[dict[str, str]], list[ValidationError]]:
    filename_lower = filename.lower()

    if filename_lower.endswith(".csv"):
        rows = read_csv(content)

    elif filename_lower.endswith(".xlsx"):
        rows = read_xlsx(content)

    else:
        raise ValueError(
            "Only CSV and XLSX files are supported."
        )

    errors = validate_columns(rows)

    if not errors:
        errors.extend(validate_rows(rows))

    return rows, errors